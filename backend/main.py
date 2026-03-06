from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import pandas as pd
import io
import asyncio
from typing import List
import shutil
import os

from .models import CrawlRequest, CrawlResult
from .crawler import LtcCrawler
from .database import init_db, get_db, SessionLocal
from .db_models import QueryRecord
from dotenv import load_dotenv
import sys
import json
from typing import Optional
from sqlalchemy import desc

load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))

# Fix for Playwright on Windows: Handled in run.py
# if sys.platform == 'win32':
#     asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

app = FastAPI()

# Global Exception Handler
from fastapi import Request
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    import traceback
    error_msg = f"GLOBAL CRASH: {str(exc)}\n{traceback.format_exc()}"
    print(error_msg, flush=True)
    with open("fatal_error.log", "a", encoding="utf-8") as f:
        f.write(error_msg + "\n" + "-"*60 + "\n")
    return JSONResponse(
        status_code=500,
        content={"message": "Internal Server Error", "detail": str(exc)},
    )

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

crawler = LtcCrawler(headless=True)

# Initialize database on startup
@app.on_event("startup")
def startup_event():
    init_db()
    print("Database initialized.")

def save_result_to_db(result_dict: dict):
    """Save a successful crawl result to the database."""
    try:
        db = SessionLocal()
        idno_masked = mask_id(result_dict.get('idno', ''))
        valid_period = result_dict.get('valid_period', '')
        
        # Check for existing record (upsert)
        existing = db.query(QueryRecord).filter_by(
            idno_masked=idno_masked,
            valid_period=valid_period
        ).first()
        
        if existing:
            record = existing
        else:
            record = QueryRecord()
            db.add(record)
        
        record.idno_masked = idno_masked
        record.name = result_dict.get('name', '')
        record.emp_id = result_dict.get('emp_id', '')
        record.organization = result_dict.get('organization', '')
        record.dob = result_dict.get('dob', '')
        record.valid_period = valid_period
        record.total_points = result_dict.get('total_points', 0.0)
        record.professional_points = result_dict.get('professional_points', 0.0)
        record.quality_points = result_dict.get('quality_points', 0.0)
        record.ethics_points = result_dict.get('ethics_points', 0.0)
        record.laws_points = result_dict.get('laws_points', 0.0)
        record.fire_safety_points = result_dict.get('fire_safety_points', 0.0)
        record.emergency_points = result_dict.get('emergency_points', 0.0)
        record.infection_points = result_dict.get('infection_points', 0.0)
        record.gender_points = result_dict.get('gender_points', 0.0)
        record.indigenous_points = result_dict.get('indigenous_points', 0.0)
        record.raw_data_json = json.dumps(result_dict.get('raw_data', {}), ensure_ascii=False)
        record.courses_json = json.dumps(result_dict.get('courses', []), ensure_ascii=False)
        record.status = result_dict.get('status', 'success')
        
        from datetime import datetime
        record.queried_at = datetime.utcnow()
        
        db.commit()
        print(f"DB: Saved record for {idno_masked} (id={record.id})")
    except Exception as e:
        print(f"DB ERROR: Failed to save result: {e}")
        db.rollback()
    finally:
        db.close()


@app.post("/api/crawl/single")
async def crawl_single(request: CrawlRequest):
    from fastapi.responses import StreamingResponse
    import json

    print(f"DEBUG: Processing request for {request.idno}")
    
    async def iter_crawl():
        try:
            api_key = os.getenv("CAPTCHA_KEY")
            if not api_key:
                 yield json.dumps({"type": "log", "message": "錯誤: CAPTCHA_KEY 未設定"}, ensure_ascii=False) + "\n"
                 yield json.dumps({"type": "error", "message": "Server Config Error"}, ensure_ascii=False) + "\n"
                 return
            
            queue = asyncio.Queue()

            # Initial logs
            await queue.put({"type": "start", "total": 100})
            await queue.put({"type": "log", "message": f"準備查詢: {request.idno} (API Key: {api_key[:4]}...)"})
            
            # Callback: Puts events into Queue
            async def progress_bridge(cur, total, msg):
                await queue.put({"type": "progress", "current": cur})
                await queue.put({"type": "log", "message": msg})

            # Worker Task
            async def worker():
                try:
                    result = await crawler.crawl_single(request.idno, request.dob, api_key, progress_callback=progress_bridge)
                    if result.status == "success":
                         await queue.put({"type": "log", "message": "查詢成功！"})
                         result_data = result.dict()
                         await queue.put({"type": "result", "data": result_data})
                         save_result_to_db(result_data)
                    else:
                         await queue.put({"type": "log", "message": f"查詢失敗: {result.error_message}"})
                         await queue.put({"type": "error", "message": result.error_message})
                except Exception as worker_e:
                     await queue.put({"type": "log", "message": f"執行錯誤: {str(worker_e)}"})
                     await queue.put({"type": "error", "message": str(worker_e)})
                finally:
                     await queue.put(None) # Sentinel

            # Start worker background task
            asyncio.create_task(worker())
            
            # Consumer Loop
            while True:
                item = await queue.get()
                if item is None:
                    break
                yield json.dumps(item, ensure_ascii=False) + "\n"
                
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            with open("backend_error.log", "a", encoding="utf-8") as f:
                f.write(f"Error in crawl_single: {str(e)}\n{error_trace}\n{'-'*60}\n")
            
            yield json.dumps({"type": "log", "message": f"系统發生錯誤: {str(e)}"}, ensure_ascii=False) + "\n"
            yield json.dumps({"type": "error", "message": str(e)}, ensure_ascii=False) + "\n"
            
    return StreamingResponse(iter_crawl(), media_type="application/x-ndjson")

@app.post("/api/crawl/batch")
async def crawl_batch(
    file: UploadFile = File(...)
):
    from fastapi.responses import StreamingResponse
    import json

    api_key = os.getenv("CAPTCHA_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="Server configuration error: CAPTCHA_KEY missing")
        
    # Read Excel (Blocking I/O but acceptable for small files)
    contents = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    async def iter_crawl():
        try:
            print(f"DEBUG: Starting iter_crawl. DF shape: {df.shape}")
            total = len(df)
            
            # Queue for collecting events from workers
            queue = asyncio.Queue()
            # Semaphore to control concurrency (e.g., 4 parallel browsers)
            sem = asyncio.Semaphore(4) 
            
            async def worker(index, row):
                async with sem:
                    # Extract data
                    emp_id = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
                    name = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ""
                    idno = str(row.iloc[2]) if pd.notna(row.iloc[2]) else ""
                    dob = str(row.iloc[3]) if pd.notna(row.iloc[3]) else ""
                    organization = ""
                    if len(row) > 4:
                        organization = str(row.iloc[4]) if pd.notna(row.iloc[4]) else ""
                    
                    # Log Start
                    await queue.put({
                        "type": "log", 
                        "message": f"[{index+1}/{total}] 正在查詢: {name} ({mask_id(idno)})..."
                    })

                    # Simple validation
                    if not idno or not dob:
                        res_dict = CrawlResult(status="failed", error_message="Missing ID or DOB", idno=idno, dob=dob).dict()
                        res_dict["emp_id"] = emp_id
                        res_dict["name"] = name
                        res_dict["organization"] = organization
                        
                        await queue.put({"type": "result", "data": res_dict})
                        await queue.put({"type": "progress_increment"})
                        return

                    # Crawl
                    try:
                        # Add random small delay to prevent exact simultaneous hits if needed
                        await asyncio.sleep(0.1) 
                        
                        res = await crawler.crawl_single(idno, dob, api_key)
                        res.emp_id = emp_id
                        res.name = name
                        res.organization = organization
                        
                        # Result
                        result_data = res.dict()
                        await queue.put({"type": "result", "data": result_data})
                        if res.status == "success":
                            save_result_to_db(result_data)
                        
                        # Log Status
                        status_msg = "成功" if res.status == "success" else f"失敗: {res.error_message}"
                        await queue.put({
                            "type": "log", 
                            "message": f"➥ {name} 查詢結果: {status_msg}"
                        })

                    except Exception as e:
                        err_res = CrawlResult(
                            idno=idno, dob=dob, emp_id=emp_id, name=name, organization=organization,
                            status="failed", error_message=str(e)
                        )
                        await queue.put({"type": "result", "data": err_res.dict()})
                        await queue.put({"type": "log", "message": f"➥ {name} 發生錯誤: {str(e)}"})
                    
                    await queue.put({"type": "progress_increment"})

            # Yield Start Event
            yield json.dumps({"type": "start", "total": total}, ensure_ascii=False) + "\n"

            # Use crawler as context manager to share browser instance
            async with crawler:
                # Launch all workers
                tasks = []
                for index, row in df.iterrows():
                    tasks.append(asyncio.create_task(worker(index, row)))
                
                # Background waiter to signal when all are done
                async def waiter():
                    await asyncio.gather(*tasks)
                    await queue.put(None) # Sentinel
                
                asyncio.create_task(waiter())

                # Consumer Loop
                completed_count = 0
                while True:
                    item = await queue.get()
                    if item is None:
                        break
                    
                    if item.get("type") == "progress_increment":
                        completed_count += 1
                        yield json.dumps({"type": "progress", "current": completed_count}, ensure_ascii=False) + "\n"
                    else:
                        yield json.dumps(item, ensure_ascii=False) + "\n"
        
        except Exception as gen_e:
            import traceback
            print(f"CRITICAL ERROR in iter_crawl: {gen_e}")
            traceback.print_exc()
            yield json.dumps({
                "type": "log", 
                "message": f"伺服器內部錯誤: {str(gen_e)}"
            }, ensure_ascii=False) + "\n"
    
    return StreamingResponse(iter_crawl(), media_type="application/x-ndjson")

@app.get("/api/template/download")
async def download_template():
    """Generate and download an Excel template for batch crawl."""
    import tempfile
    
    # Create template with correct headers and example data
    template_data = {
        '員工編號': ['EMP001', 'EMP002'],
        '姓名': ['王小明', '李小華'],
        '身分證字號': ['A123456789', 'B234567890'],
        '出生年月日': ['080/01/15', '075/06/20'],
        '機構': ['XX護理之家', 'OO長照中心'],
    }
    df = pd.DataFrame(template_data)
    
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp_path = tmp.name
    tmp.close()
    
    with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='名冊')
        
        # Auto-adjust column widths
        ws = writer.sheets['名冊']
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)), *(len(str(v)) for v in df[col_name]))
            ws.column_dimensions[chr(64 + col_idx)].width = max_len * 2.5 + 4

    return FileResponse(
        tmp_path,
        filename="批次查詢範本.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.post("/api/export/excel")
async def export_excel(request: Request):
    """Export query results to a styled Excel file."""
    import tempfile
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    body = await request.json()
    results_data = body.get("results", [])
    
    if not results_data:
        raise HTTPException(status_code=400, detail="No results to export")
    
    # Build main summary rows
    rows = []
    for r in results_data:
        raw = r.get("raw_data", {})
        rows.append({
            '員工編號': r.get('emp_id', ''),
            '姓名': r.get('name', ''),
            '機構': r.get('organization', ''),
            '身分證字號': mask_id(r.get('idno', '')),
            '有效期限': r.get('valid_period', ''),
            '狀態': '成功' if r.get('status') == 'success' else '失敗',
            '專業課程(實體)': raw.get('prof_physical', 0),
            '專業課程(網路)': raw.get('prof_online', 0),
            '專業品質(實體)': raw.get('qual_physical', 0),
            '專業品質(網路)': raw.get('qual_online', 0),
            '專業倫理(實體)': raw.get('ethic_physical', 0),
            '專業倫理(網路)': raw.get('ethic_online', 0),
            '專業法規(實體)': raw.get('law_physical', 0),
            '專業法規(網路)': raw.get('law_online', 0),
            '消防安全': raw.get('fire_safety', 0),
            '緊急應變': raw.get('emergency', 0),
            '感染管制': raw.get('infection', 0),
            '性別敏感度': raw.get('gender', 0),
            '原住民族(舊)': raw.get('indigenous_legacy', 0),
            '原住民族文化': raw.get('indigenous_culture', 0),
            '多元族群文化': raw.get('diverse_culture', 0),
            '實體總積分': raw.get('total_physical', 0),
            '網路(舊)': raw.get('total_online_old', 0),
            '網路(新)': raw.get('total_online_new', 0),
            '總積分': r.get('total_points', 0),
        })
    
    df = pd.DataFrame(rows)
    
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp_path = tmp.name
    tmp.close()
    
    with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='查詢結果')
        
        ws = writer.sheets['查詢結果']
        
        # Style: Header row
        header_fill = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        thin_border = Border(
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Style: Data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
        
        # Auto-adjust column widths
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            max_len = max(len(str(cell.value or '')) for cell in col_cells)
            ws.column_dimensions[col_letter].width = max(max_len * 2 + 4, 10)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # --- Sheet 2: Course details (if any) ---
        course_rows = []
        for r in results_data:
            person_name = r.get('name', '') or r.get('idno', '')
            for c in r.get('courses', []):
                course_rows.append({
                    '姓名': person_name,
                    '日期': c.get('date', ''),
                    '課程名稱': c.get('name', ''),
                    '實施方式': c.get('mode', ''),
                    '開課單位': c.get('unit', ''),
                    '課程屬性': c.get('attribute', ''),
                    '課程類別': c.get('category', ''),
                    '訓練課程': c.get('training_course', ''),
                    '積分': c.get('points', 0),
                    '認可狀態': c.get('status', ''),
                })
        
        if course_rows:
            df2 = pd.DataFrame(course_rows)
            df2.to_excel(writer, index=False, sheet_name='課程明細')
            
            ws2 = writer.sheets['課程明細']
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for col_cells in ws2.columns:
                col_letter = col_cells[0].column_letter
                max_len = max(len(str(cell.value or '')) for cell in col_cells)
                ws2.column_dimensions[col_letter].width = max(max_len * 2 + 4, 10)
            ws2.freeze_panes = 'A2'
    
    return FileResponse(
        tmp_path,
        filename="長照積分查詢結果.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def mask_id(idno):
    if len(idno) > 6:
        return idno[:3] + "***" + idno[-3:]
    return idno

# =================== History API ===================

@app.get("/api/history")
def get_history(
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 50
):
    """Get query history with optional search and pagination."""
    db = SessionLocal()
    try:
        query = db.query(QueryRecord).order_by(desc(QueryRecord.queried_at))
        
        if search:
            pattern = f"%{search}%"
            query = query.filter(
                (QueryRecord.name.ilike(pattern)) |
                (QueryRecord.idno_masked.ilike(pattern)) |
                (QueryRecord.organization.ilike(pattern)) |
                (QueryRecord.emp_id.ilike(pattern))
            )
        
        total = query.count()
        records = query.offset((page - 1) * page_size).limit(page_size).all()
        
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "results": [r.to_dict() for r in records]
        }
    finally:
        db.close()

@app.get("/api/history/{record_id}")
def get_history_detail(record_id: int):
    """Get a single history record by ID."""
    db = SessionLocal()
    try:
        record = db.query(QueryRecord).filter_by(id=record_id).first()
        if not record:
            raise HTTPException(status_code=404, detail="Record not found")
        return record.to_dict()
    finally:
        db.close()

@app.delete("/api/history/{record_id}")
def delete_history_record(record_id: int):
    """Delete a single history record."""
    db = SessionLocal()
    try:
        record = db.query(QueryRecord).filter_by(id=record_id).first()
        if not record:
            raise HTTPException(status_code=404, detail="Record not found")
        db.delete(record)
        db.commit()
        return {"message": "Deleted", "id": record_id}
    finally:
        db.close()

# =================== Static Files (MUST be last) ===================
# app.mount("/") catches ALL unmatched routes, so it must come AFTER all API routes.
frontend_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend", "dist")
if os.path.exists(frontend_path):
    app.mount("/", StaticFiles(directory=frontend_path, html=True), name="static")
else:
    @app.get("/")
    def read_root():
        return {"status": "ok", "service": "LTC Credit Crawler (Static files not found)"}
